"""utils/export.py

Multi-sheet Excel export engine for ECC Micromechanics Calculator.

Responsibilities:
  - build_summary_df        → Sheet 1: long-format results table
  - build_sigma_delta_df    → Sheet 2: wide-format σ–δ curve matrix
  - build_settings_log_df   → Sheet 3: raw inputs for reproducibility
  - DataExportWorker        → QThread that owns all I/O, never touches the main thread
"""
from __future__ import annotations

import logging
from pathlib import Path

import numpy as np
import pandas as pd
from PySide6.QtCore import QThread, Signal

from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from core.engine import AnalysisResult
from models.project import ProjectModel, SeriesEntry

_log = logging.getLogger(__name__)

_SIM_PARAM_FIELDS: list[tuple[str, str]] = [
    ("sim_fiber_type", "Fiber Type"),
    ("sim_V_f", "V_f (vol. fraction)"),
    ("sim_L_f", "L_f (mm)"),
    ("sim_E_f", "E_f (GPa)"),
    ("sim_sigma_fu", "sigma_fu (MPa)"),
    ("sim_G_d", "G_d (J/m^2)"),
    ("sim_beta", "beta (slip-hardening)"),
    ("sim_f_snubbing", "f_snubbing"),
    ("sim_n_delta_points", "Curve Points"),
    ("sim_P_anchor_max", "P_anchor_max (N)"),
    ("sim_delta_hook", "delta_hook (mm)"),
]

# ---------------------------------------------------------------------------
# Pure data-transformation functions (zero Qt dependency, fully testable)
# ---------------------------------------------------------------------------

def build_summary_df(model: ProjectModel) -> pd.DataFrame:
    rows: list[dict] = []

    for entry in model:
        if entry.result is None:
            continue

        r: AnalysisResult = entry.result
        p = entry.params

        rows.append(
            {
                "Series Name":       p.name,
                "Variable Name":     model.variable_name,
                "Variable Value":    round(p.variable_value, 3),
                "tau0 (MPa)":        round(r.tau0, 3),
                "E_m (GPa)":         round(p.e_m, 3),
                "K_m (MPa*m^0.5)":   round(r.km, 3),
                "sigma_fc (MPa)":    round(p.sigma_fc, 3),
                "J_tip (J/m^2)":     round(r.j_tip, 3),
                "sigma0 (MPa)":      round(r.sigma0, 3),
                "delta0 (mm)":       round(r.delta0, 3),
                "J_b' (J/m^2)":      round(r.jb_prime, 3),
                "PSH Strength":      round(r.psh_strength, 3),
                "PSH Energy":        round(r.psh_energy, 3),
            }
        )

    if not rows:
        return pd.DataFrame(
            columns=[
                "Series Name", "Variable Name", "Variable Value",
                "tau0 (MPa)", "E_m (GPa)", "K_m (MPa*m^0.5)",
                "sigma_fc (MPa)", "J_tip (J/m^2)", "sigma0 (MPa)",
                "delta0 (mm)", "J_b' (J/m^2)", "PSH Strength", "PSH Energy",
            ]
        )

    return pd.DataFrame(rows)


def build_sigma_delta_df(model: ProjectModel) -> pd.DataFrame:
    column_pairs: list[pd.DataFrame] = []

    for entry in model:
        df = entry.params.sigma_delta_df
        if df is None or df.empty:
            continue

        name = entry.params.name
        pair = pd.DataFrame(
            {
                f"{name}_delta (mm)": df["delta"].reset_index(drop=True),
                f"{name}_sigma (MPa)": df["sigma"].reset_index(drop=True),
            }
        )
        column_pairs.append(pair)

    if not column_pairs:
        return pd.DataFrame()

    return pd.concat(column_pairs, axis=1)


def build_settings_log_df(model: ProjectModel) -> pd.DataFrame:
    rows: list[dict] = []

    for entry in model:
        p = entry.params
        data_source = {
            "csv": "Imported CSV",
            "simulation": "Theoretical Simulation",
        }.get(p.sigma_delta_source, "None")

        row: dict = {
            "Series Name":    p.name,
            "Data Source":    data_source,
            "P_peak (N)":     p.p_peak,
            "d_f (mm)":       p.d_f,
            "L_e (mm)":       p.l_e,
            "P_max (N)":      p.p_max,
            "Span S (mm)":    p.span,
            "Width b (mm)":   p.b,
            "Depth d (mm)":   p.d,
            "Notch a0 (mm)":  p.a0,
            "E_m (GPa)":      p.e_m,
            "sigma_fc (MPa)": p.sigma_fc,
        }

        for attr, col in _SIM_PARAM_FIELDS:
            if hasattr(p, attr):
                row[col] = getattr(p, attr)
            else:
                row[col] = np.nan
                _log.warning(
                    "build_settings_log_df: SeriesParams has no attribute %r; "
                    "column %r will be NaN. Check _SIM_PARAM_FIELDS.",
                    attr, col,
                )

        rows.append(row)

    if not rows:
        sim_cols = [col for _, col in _SIM_PARAM_FIELDS]
        return pd.DataFrame(
            columns=[
                "Series Name", "Data Source",
                "P_peak (N)", "d_f (mm)", "L_e (mm)", "P_max (N)",
                "Span S (mm)", "Width b (mm)", "Depth d (mm)", "Notch a0 (mm)",
                "E_m (GPa)", "sigma_fc (MPa)",
                *sim_cols,
            ]
        )

    return pd.DataFrame(rows)


def write_excel(
    path: Path,
    summary_df: pd.DataFrame,
    sigma_delta_df: pd.DataFrame,
    settings_df: pd.DataFrame,
) -> None:
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        summary_df.to_excel(writer, sheet_name="Summary_Results", index=False)
        sigma_delta_df.to_excel(
            writer, sheet_name="Sigma_Delta_Curves", index=False
        )
        settings_df.to_excel(
            writer, sheet_name="Project_Settings_Log", index=False
        )

        workbook = writer.book

        pass_fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
        fail_fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")

        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]

            worksheet.freeze_panes = "A2"

            for cell in worksheet[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")

            for col in worksheet.columns:
                max_length = 0
                col_letter = col[0].column_letter

                for cell in col:
                    try:
                        if cell.value is not None:
                            max_length = max(max_length, len(str(cell.value)))
                    except Exception:
                        pass

                worksheet.column_dimensions[col_letter].width = max_length + 2.5

            if sheet_name == "Summary_Results":
                headers = {cell.value: cell.column for cell in worksheet[1]}

                strength_col = headers.get("PSH Strength")
                energy_col = headers.get("PSH Energy")

                for row in range(2, worksheet.max_row + 1):
                    if strength_col:
                        cell = worksheet.cell(row=row, column=strength_col)
                        try:
                            val = float(cell.value)
                            cell.fill = pass_fill if val >= AnalysisResult.PSH_STRENGTH_THRESHOLD else fail_fill
                        except (TypeError, ValueError):
                            pass

                    if energy_col:
                        cell = worksheet.cell(row=row, column=energy_col)
                        try:
                            val = float(cell.value)
                            cell.fill = pass_fill if val >= AnalysisResult.PSH_ENERGY_THRESHOLD else fail_fill
                        except (TypeError, ValueError):
                            pass


# ---------------------------------------------------------------------------
# QThread worker — owns all disk I/O, never runs in the main thread
# ---------------------------------------------------------------------------

class DataExportWorker(QThread):
    progress = Signal(int, str)
    finished = Signal(object)
    failed = Signal(str)

    def __init__(self, model: ProjectModel, output_path: Path) -> None:
        super().__init__()
        self._summary_df = build_summary_df(model)
        self._sigma_delta_df = build_sigma_delta_df(model)
        self._settings_df = build_settings_log_df(model)
        self._output_path = output_path.resolve()

    def run(self) -> None:
        try:
            self.progress.emit(20, "Writing Summary_Results…")
            self.progress.emit(50, "Writing Sigma_Delta_Curves…")
            self.progress.emit(75, "Writing Project_Settings_Log…")
            write_excel(
                self._output_path,
                self._summary_df,
                self._sigma_delta_df,
                self._settings_df,
            )

        except PermissionError:
            self.failed.emit(
                f"Export failed: '{self._output_path.name}' is open in another "
                f"application. Please close it and try again."
            )
            return
        except OSError as exc:
            self.failed.emit(f"Export failed (I/O error): {exc}")
            return
        except Exception as exc:  # noqa: BLE001
            self.failed.emit(f"Export failed (unexpected): {exc}")
            return

        self.progress.emit(100, "Done.")
        self.finished.emit(self._output_path)
